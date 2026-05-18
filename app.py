import streamlit as st
from docxtpl import DocxTemplate
import io
import os
import pandas as pd
from datetime import datetime

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Smart Security | Portal", page_icon="💳", layout="centered")

# --- ESTILOS CORPORATIVOS ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght=700&display=swap');
    
    /* OCULTAR MENÚS DE STREAMLIT */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display:none;}
    
    /* FONDO DE LA APLICACIÓN */
    .stApp { 
        background: radial-gradient(circle at 20% 30%, #003a85 0%, #001B3D 60%, #FF7F00 130%) !important; 
        background-attachment: fixed; 
    }
    
    /* --- CORRECCIÓN DE COLOR PARA TEXTOS SOBRE EL FONDO AZUL --- */
    /* Cambia a blanco el texto del selector principal y de las preguntas que están afuera */
    .stSelectbox label p, .stNumberInput label p, .stRadio label p, p {
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 600;
    }
    
    /* --- CONTENEDOR DE LOS FORMULARIOS (NO SE CAMBIA, SE MANTIENE BLANCO) --- */
    [data-testid="stForm"] { 
        background-color: #FFFFFF !important; 
        border-radius: 20px !important; 
        padding: 25px !important; 
        box-shadow: 0 15px 35px rgba(0,0,0,0.4); 
    }
    
    /* Textos internos del formulario se mantienen oscuros para que contrasten con el fondo blanco */
    [data-testid="stForm"] label p, [data-testid="stForm"] h2, [data-testid="stForm"] h3 { 
        color: #001B3D !important; 
        font-family: 'Montserrat', sans-serif !important; 
        font-weight: 700; 
    }
    [data-testid="stForm"] input { color: #000000 !important; }
    
    /* Botón del Formulario */
    [data-testid="stForm"] button { 
        background-color: #001B3D !important; 
        color: white !important; 
        font-weight: 800; 
        width: 100%; 
        border: 2px solid #FF7F00; 
        border-radius: 12px; 
        padding: 12px;
    }
    @media (max-width: 640px) {
        [data-testid="stForm"] { padding: 15px !important; }
    }
    </style>
    """, unsafe_allow_html=True)

# --- LOGO ---
if os.path.exists("hunter1.png"):
    col_l, col_c, col_r = st.columns([1.5, 1, 1.5])
    with col_c: st.image("hunter1.png", width=150)

# --- SELECTOR PRINCIPAL ---
categoria = st.selectbox("📂 Tipo de Documento", ["Declaración Jurada", "Contrato de Alianza Comercial", "Formato Creación Usuarios"])

# =========================================================================
# VISTA 1: FORMULARIO DINÁMICO PARA EXCEL DE CREACIÓN DE USUARIOS
# =========================================================================
if categoria == "Formato Creación Usuarios":
    with st.form("form_creacion_usuarios"):
        st.markdown("<h2 style='text-align:center;'>📝 Registro de Usuarios</h2>", unsafe_allow_html=True)
        st.markdown("### 🏢 Datos Generales de la Empresa")
        
        col1, col2 = st.columns(2)
        with col1:
            rep_legal = st.text_input("Nombre del Representante Legal")
            razon_social = st.text_input("Nombre de Tienda / Razón Social")
            ruc = st.text_input("RUC")
            correo = st.text_input("Correo Electrónico")
        with col2:
            telefono1 = st.text_input("Número Celular 1")
            telefono2 = st.text_input("Número Celular 2")
            banco = st.text_input("Banco (ej. BBVA, BCP)")
            tipo_cuenta = st.text_input("Tipo de Cuenta", value="AHORROS")
            n_cuenta = st.text_input("Número de Cuenta + CCI")

        st.markdown("<hr style='border: 0.5px solid #001B3D;'>", unsafe_allow_html=True)
        
        # --- SELECCIÓN DINÁMICA DE CANTIDADES ---
        st.markdown("### ⚙️ Configuración de Filas")
        num_tiendas = st.number_input("¿Cuántas tiendas desea registrar?", min_value=1, max_value=15, value=1, step=1)
        num_usuarios = st.number_input("¿Cuántos usuarios / vendedores desea registrar?", min_value=1, max_value=50, value=1, step=1)

        # --- SECCIÓN: TIENDAS DINÁMICAS ---
        st.markdown("<hr style='border: 0.5px solid #001B3D;'>", unsafe_allow_html=True)
        st.markdown("### 📍 Relación de Tiendas")
        lista_tiendas = []
        for i in range(int(num_tiendas)):
            st.markdown(f"**🏪 Tienda {i+1}**")
            t_col1, t_col2 = st.columns(2)
            with t_col1:
                t_nom = st.text_input(f"Nombre Tienda {i+1}", key=f"t_nom_{i}")
                t_dir = st.text_input(f"Dirección Tienda {i+1}", key=f"t_dir_{i}")
            with t_col2:
                t_dep = st.text_input(f"Departamento Tienda {i+1}", value="LIMA", key=f"t_dep_{i}")
                t_ciu = st.text_input(f"Ciudad / Distrito Tienda {i+1}", key=f"t_ciu_{i}")
            lista_tiendas.append({"nombre": t_nom, "direccion": t_dir, "departamento": t_dep, "ciudad": t_ciu})

        # --- SECCIÓN: USUARIOS DINÁMICOS ---
        st.markdown("<hr style='border: 0.5px solid #001B3D;'>", unsafe_allow_html=True)
        st.markdown("### 👥 Relación de Usuarios / Vendedores")
        lista_usuarios = []
        for j in range(int(num_usuarios)):
            st.markdown(f"**👤 Vendedor {j+1}**")
            v_col1, v_col2 = st.columns(2)
            with v_col1:
                v_tnd = st.text_input(f"Tienda Asignada Vendedor {j+1}", key=f"v_tnd_{j}")
                v_nom = st.text_input(f"Nombre Completo Vendedor {j+1}", key=f"v_nom_{j}")
            with v_col2:
                v_crr = st.text_input(f"Correo Vendedor {j+1}", key=f"v_crr_{j}")
                v_cel = st.text_input(f"Celular Vendedor {j+1}", key=f"v_cel_{j}")
            lista_usuarios.append({"tienda": v_tnd, "nombre": v_nom, "correo": v_crr, "celular": v_cel})

        st.markdown("<br>", unsafe_allow_html=True)
        submit_u = st.form_submit_button("🚀 GENERAR EXCEL DE USUARIOS")

    if submit_u:
        try:
            import openpyxl
            
            archivo_plantilla = "USUARIOS.xlsx"
            if not os.path.exists(archivo_plantilla):
                st.error(f"❌ No se encontró el archivo '{archivo_plantilla}' en tu GitHub. Por favor, súbelo.")
                st.stop()
                
            wb = openpyxl.load_workbook(archivo_plantilla)
            ws = wb["ORIENTE SMART"]
            
            # --- DETECTOR INTELIGENTE DE CELDAS DE RESPUESTA ---
            fila_inicio_tiendas = None
            fila_inicio_vendedores = None
            
            for row in range(1, 100):
                for col in range(1, 6):
                    celda_val = str(ws.cell(row=row, column=col).value).strip().upper()
                    
                    if "REPRESENTANTE" in celda_val:
                        ws.cell(row=row, column=col+1, value=rep_legal.upper())
                    elif "NOMBRE DE TIENDA" in celda_val or "RAZÓN SOCIAL" in celda_val:
                        ws.cell(row=row, column=col+1, value=razon_social.upper())
                    elif "RUC" in celda_val:
                        ws.cell(row=row, column=col+1, value=ruc)
                    elif "CORREO" in celda_val and row < 20:
                        ws.cell(row=row, column=col+1, value=correo.upper())
                    elif "CELULAR 1" in celda_val or "TELEFONO 1" in celda_val:
                        ws.cell(row=row, column=col+1, value=telefono1)
                    elif "CELULAR 2" in celda_val or "TELEFONO 2" in celda_val:
                        ws.cell(row=row, column=col+1, value=telefono2)
                    elif "BANCO" in celda_val:
                        ws.cell(row=row, column=col+1, value=banco.upper())
                    elif "TIPO DE CUENTA" in celda_val:
                        ws.cell(row=row, column=col+1, value=tipo_cuenta.upper())
                    elif "NUMERO DE CUENTA" in celda_val or "NÚMERO DE CUENTA" in celda_val:
                        ws.cell(row=row, column=col+1, value=n_cuenta)
                        
                    if "NOMBRE TIENDA" in celda_val and fila_inicio_tiendas is None:
                        fila_inicio_tiendas = row + 1
                    elif "NOMBRE TIENDA" in celda_val and fila_inicio_tiendas is not None:
                        fila_inicio_vendedores = row + 1

            if not fila_inicio_tiendas: fila_inicio_tiendas = 22
            if not fila_inicio_vendedores: fila_inicio_vendedores = 42

            for tienda in lista_tiendas:
                if tienda["nombre"]:
                    ws.cell(row=fila_inicio_tiendas, column=2, value=tienda["nombre"].upper())
                    ws.cell(row=fila_inicio_tiendas, column=3, value=tienda["direccion"].upper())
                    ws.cell(row=fila_inicio_tiendas, column=4, value=tienda["departamento"].upper())
                    ws.cell(row=fila_inicio_tiendas, column=5, value=tienda["ciudad"].upper())
                    fila_inicio_tiendas += 1
            
            for usuario in lista_usuarios:
                if usuario["nombre"]:
                    ws.cell(row=fila_inicio_vendedores, column=2, value=usuario["tienda"].upper())
                    ws.cell(row=fila_inicio_vendedores, column=3, value=usuario["nombre"].upper())
                    ws.cell(row=fila_inicio_vendedores, column=4, value=usuario["correo"])
                    ws.cell(row=fila_inicio_vendedores, column=5, value=usuario["celular"])
                    fila_inicio_vendedores += 1
            
            output_excel = io.BytesIO()
            wb.save(output_excel)
            output_excel.seek(0)
            
            st.balloons()
            st.success("✅ ¡Formato de Creación de Usuarios generado con éxito!")
            nombre_file = razon_social.replace(" ", "_") if razon_social else "Usuarios"
            st.download_button(
                label="📊 DESCARGAR EXCEL DE USUARIOS", 
                data=output_excel, 
                file_name=f"Usuarios_{nombre_file}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Ocurrió un error al procesar la plantilla de Excel: {e}")

# =========================================================================
# VISTA 2: FORMULARIOS ANTERIORES (CONTRATOS Y DJ TRADICIONALES)
# =========================================================================
else:
    tipo_persona = st.radio("👤 Perfil", ["Natural", "Jurídica"], horizontal=True)
    
    with st.form("form_documentos_tradicionales"):
        st.markdown("<h2 style='text-align:center;'>📝 Registro de Información</h2>", unsafe_allow_html=True)
        
        if tipo_persona == "Natural":
            col1, col2 = st.columns(2)
            with col1:
                nombre = st.text_input("Nombres y Apellidos")
                direccion = st.text_input("Dirección / Domicilio Declarado")
                correo = st.text_input("Correo Electrónico")
                pais = st.text_input("País de Origen/Residencia", value="PERÚ")
            with col2:
                documento = st.text_input("DNI / CE")
                ruc_natural = st.text_input("RUC (Opcional, necesario para Contrato)")
                telefono = st.text_input("Teléfono / Celular")
                ciudad = st.text_input("Ciudad de Firma", value="Lima")
                
        else:
            st.markdown("### 👤 Datos del Representante Legal")
            r1c1, r1c2 = st.columns(2)
            with r1c1:
                rep_legal_old = st.text_input("Nombres y Apellidos (Representante)")
                dni_rep = st.text_input("DNI del Representante")
                fecha_nac_rep = st.text_input("Fecha de Nacimiento (DD/MM/AAAA)")
            with r1c2:
                nacionalidad_rep = st.text_input("Nacionalidad", value="PERUANA")
                correo_rep = st.text_input("Correo Electrónico")
                tel_rep = st.text_input("Teléfono de Contacto")

            st.markdown("<hr style='border: 0.5px solid #001B3D;'>", unsafe_allow_html=True)
            st.markdown("### 🏢 Datos de la Empresa")
            r2c1, r2c2 = st.columns(2)
            with r2c1:
                razon_social_old = st.text_input("Razón Social")
                ruc_old = st.text_input("RUC")
                partida = st.text_input("N° de Partida Registral")
            with r2c2:
                direccion_emp = st.text_input("Dirección Fiscal")
                asiento = st.text_input("N° de Asiento")
                ciudad_f = st.text_input("Ciudad de Firma", value="Lima")

        st.markdown("<br>", unsafe_allow_html=True)
        submit_doc = st.form_submit_button("🚀 GENERAR DOCUMENTO OFICIAL")

    if submit_doc:
        try:
            contexto = {}
            datos_excel = {}
            nombre_para_archivo = "Documento"
            
            if tipo_persona == "Natural":
                if categoria == "Declaración Jurada":
                    contexto = {
                        "nombres_apellidos": nombre, "numero_documento": documento,
                        "dirección_declarada": direccion, "direccion_declarada": direccion,     
                        "dirección": direccion, "direccion": direccion,                
                        "numero_telefono": telefono, "telefono": telefono,
                        "correo_electronico": correo, "ciudad": ciudad, "pais": pais, "dni_x": "X"
                    }
                else:
                    contexto = {
                        "nombre_persona_natural": nombre, "direccion": direccion, "dirección": direccion,
                        "dirección_declarada": direccion, "direccion_declarada": direccion,
                        "numero_ruc": ruc_natural, "numero_dni": documento,
                        "numero_telefono": telefono, "telefono": telefono,
                        "correo_electronico": correo, "ciudad": ciudad, "pais": pais
                    }
                datos_excel = {
                    "Fecha Registro": [datetime.now().strftime("%d/%m/%Y")], "Tipo Documento": [categoria],
                    "Perfil": ["Natural"], "Nombre / Razón Social": [nombre], "DNI / CE": [documento],
                    "RUC": [ruc_natural], "Dirección": [direccion], "Teléfono": [telefono],
                    "Correo": [correo], "Ciudad": [ciudad]
                }
                nombre_para_archivo = nombre.replace(" ", "_") if nombre else "Natural"
            else:
                contexto = {
                    "nombre_persona_natural": rep_legal_old, 
                    "numero_dni": dni_rep,
                    "fecha_texto_nacimiento": fecha_nac_rep, 
                    "nacionalidad": nacionalidad_rep,
                    "correo_electronico": correo_rep, 
                    "numero_telefono": tel_rep, 
                    "telefono": tel_rep,                   
                    "razon_social": razon_social_old, 
                    "numero_ruc": ruc_old,                     
                    "dirección": direccion_emp, 
                    "direccion": direccion_emp,              
                    "dirección_declarada": direccion_emp, 
                    "direccion_declarada": direccion_emp,
                    "numero_partida_registral": partida, 
                    "numero_asiento": asiento,
                    "ciudad": ciudad_f, 
                    "pais": "PERÚ", 
                    "dni_x": "X",
                    "pas_x": " ", 
                    "ce_x": " ", 
                    "sol_x": " ", 
                    "cas_x": " ", 
                    "div_x": " ", 
                    "viu_x": " ", 
                    "con_x": " "
                }
                datos_excel = {
                    "Fecha Registro": [datetime.now().strftime("%d/%m/%Y")], "Tipo Documento": [categoria],
                    "Perfil": ["Jurídica"], "Nombre / Razón Social": [razon_social_old], "DNI / CE": [dni_rep],
                    "RUC": [ruc_old], "Dirección": [direccion_emp], "Teléfono": [tel_rep],
                    "Correo": [correo_rep], "Ciudad": [ciudad_f], "Representante Legal": [rep_legal_old],
                    "Partida Registral": [partida], "Asiento": [asiento]
                }
                nombre_para_archivo = razon_social_old.replace(" ", "_") if razon_social_old else "Juridica"

            if categoria == "Declaración Jurada":
                archivo = "djpersonajuridica.docx" if tipo_persona == "Jurídica" else "Djnatural.docx"
            else:
                archivo = "contratojuridica.docx" if tipo_persona == "Jurídica" else "contratonatural.docx"
                
            doc = DocxTemplate(archivo)
            hoy = datetime.now()
            meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
            contexto["fecha_texto"] = f"{hoy.day} de {meses[hoy.month - 1]} de {hoy.year}"
            
            doc.render(contexto)

            if tipo_persona == "Jurídica":
                for table in doc.tables:
                    for row in table.rows:
                        for cell in row.cells:
                            if "11641837" in cell.text:
                                cell.text = cell.text.replace("11641837", contexto["numero_partida_registral"])

            output_word = io.BytesIO()
            doc.save(output_word)
            output_word.seek(0)
            
            output_excel = io.BytesIO()
            df = pd.DataFrame(datos_excel)
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Registro")
            output_excel.seek(0)
            
            st.balloons()
            st.success("✅ ¡Documentos estructurados con éxito!")
            
            tipo_doc_nombre = "DJ" if categoria == "Declaración Jurada" else "Contrato"
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                st.download_button(
                    label="📥 DESCARGAR WORD", data=output_word, 
                    file_name=f"{tipo_doc_nombre}_{nombre_para_archivo}.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                )
            with col_btn2:
                st.download_button(
                    label="📊 DESCARGAR EXCEL", data=output_excel, 
                    file_name=f"Reporte_{tipo_doc_nombre}_{nombre_para_archivo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Error: Revisa que tus plantillas base .docx estén cargadas en GitHub.")
            st.info(f"Detalle técnico: {e}")

st.markdown("<p style='text-align: center; color: white; font-size: 12px; margin-top: 50px;'>Willy Ríos | Smart Security © 2026</p>", unsafe_allow_html=True)
